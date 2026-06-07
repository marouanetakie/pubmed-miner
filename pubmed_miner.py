"""
PubMed Miner v3.0
Fetch, score, cache, and export biomedical literature from PubMed and Europe PMC.

New in v3.0: Europe PMC dual-source search, SQLite result cache, Claude AI structured
extraction from abstracts, BibTeX/RIS reference export, Excel publication charts.

Author:      Marouane Takie
Affiliation: PhD Candidate | Physiology, Pharmacology & Phytochemistry
Institution: Université Sidi Mohamed Ben Abdellah (USMBA), Fès, Morocco
ORCID:       0009-0009-8621-8548
GitHub:      github.com/marouanetakie
"""

import argparse
import json
import os
import sqlite3
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd

try:
    import anthropic as _anthropic
    _ANTHROPIC_OK = True
except ImportError:
    _ANTHROPIC_OK = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

NCBI_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
EPMC_BASE = "https://www.ebi.ac.uk/europepmc/webservices/rest/"
AI_MODEL  = "claude-haiku-4-5"

DEFAULT_QUERIES = [
    "Erodium moschatum",
    "Reseda alba",
    "Moroccan medicinal plants antidiabetic",
    "silver nanoparticles green synthesis antioxidant",
    "UPLC-ESI-MS plant extract phenolic compounds",
]

BATCH_SIZE = 20

COLUMNS_BASE = [
    "UID", "Source", "PMID", "Queries", "Title", "Authors",
    "Journal", "Year", "DOI", "Relevance_Score",
    "Abstract", "Keywords", "MeSH_Terms",
]
COLUMNS_AI = [
    "AI_Species", "AI_Compounds", "AI_Activities",
    "AI_Key_Results", "AI_Study_Type",
]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "PubMed Miner v3.0 — multi-source fetch, SQLite cache, AI extraction, "
            "BibTeX/RIS export, Excel charts."
        )
    )
    p.add_argument(
        "--queries", nargs="+", default=None, metavar="QUERY",
        help=(
            "One or more search queries. "
            "Defaults to 5 built-in phytochemistry queries."
        ),
    )
    p.add_argument(
        "--source", choices=["pubmed", "europepmc", "both"], default="pubmed",
        help="Literature source (default: pubmed).",
    )
    p.add_argument(
        "--max-results", type=int, default=50, metavar="N",
        help="Maximum results per query per source (default: 50).",
    )
    p.add_argument(
        "--year-from", type=int, default=None, metavar="YEAR",
        help="Filter to articles published from YEAR onwards.",
    )
    p.add_argument(
        "--relevance-threshold", type=int, default=60, metavar="N",
        help="Min relevance score for the High Relevance sheet (default: 60).",
    )
    p.add_argument(
        "--ai-extract", action="store_true",
        help=(
            "Use Claude AI to extract species, compounds, activities, and key "
            "results from each abstract. Requires ANTHROPIC_API_KEY env var."
        ),
    )
    p.add_argument(
        "--api-key", default=None, metavar="KEY",
        help="NCBI API key (raises rate limit from 3 to 10 req/s).",
    )
    p.add_argument(
        "--cache-db", default="pubmed_cache.db", metavar="FILE",
        help="SQLite cache file (default: pubmed_cache.db).",
    )
    p.add_argument(
        "--no-cache", action="store_true",
        help="Bypass the SQLite cache and re-fetch all records.",
    )
    p.add_argument(
        "--output", default=None, metavar="FILE",
        help="Output .xlsx filename (default: pubmed_results_YYYYMMDD_HHMMSS.xlsx).",
    )
    p.add_argument(
        "--export-bibtex", default=None, metavar="FILE",
        help="Also export results as BibTeX (.bib).",
    )
    p.add_argument(
        "--export-ris", default=None, metavar="FILE",
        help="Also export results as RIS (.ris).",
    )
    return p.parse_args()


# ---------------------------------------------------------------------------
# SQLite cache
# ---------------------------------------------------------------------------

class ArticleCache:
    def __init__(self, db_path: str, bypass: bool = False):
        self.bypass = bypass
        self.conn = sqlite3.connect(db_path) if not bypass else None
        if not bypass:
            self._setup()

    def _setup(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS articles (
                uid         TEXT PRIMARY KEY,
                source      TEXT,
                record_json TEXT,
                cached_at   TEXT
            );
            CREATE TABLE IF NOT EXISTS ai_extractions (
                uid             TEXT PRIMARY KEY,
                extraction_json TEXT,
                model           TEXT,
                extracted_at    TEXT
            );
        """)
        self.conn.commit()

    def get(self, uid: str) -> dict | None:
        if self.bypass or self.conn is None:
            return None
        row = self.conn.execute(
            "SELECT record_json FROM articles WHERE uid = ?", (uid,)
        ).fetchone()
        return json.loads(row[0]) if row else None

    def put(self, uid: str, source: str, record: dict):
        if self.bypass or self.conn is None:
            return
        self.conn.execute(
            "INSERT OR REPLACE INTO articles VALUES (?, ?, ?, ?)",
            (uid, source, json.dumps(record), datetime.now().isoformat()),
        )
        self.conn.commit()

    def get_ai(self, uid: str) -> dict | None:
        if self.bypass or self.conn is None:
            return None
        row = self.conn.execute(
            "SELECT extraction_json FROM ai_extractions WHERE uid = ?", (uid,)
        ).fetchone()
        return json.loads(row[0]) if row else None

    def put_ai(self, uid: str, extraction: dict, model: str):
        if self.bypass or self.conn is None:
            return
        self.conn.execute(
            "INSERT OR REPLACE INTO ai_extractions VALUES (?, ?, ?, ?)",
            (uid, json.dumps(extraction), model, datetime.now().isoformat()),
        )
        self.conn.commit()

    def close(self):
        if self.conn:
            self.conn.close()


# ---------------------------------------------------------------------------
# PubMed source
# ---------------------------------------------------------------------------

def search_pubmed(
    query: str, max_results: int, year_from: int | None, api_key: str | None
) -> list[str]:
    params: dict = {
        "db": "pubmed", "term": query,
        "retmax": max_results, "retmode": "json", "sort": "relevance",
    }
    if year_from:
        params.update({"datetype": "pdat", "mindate": str(year_from), "maxdate": "3000"})
    if api_key:
        params["api_key"] = api_key
    r = requests.get(NCBI_BASE + "esearch.fcgi", params=params, timeout=30)
    r.raise_for_status()
    return r.json()["esearchresult"]["idlist"]


def fetch_details(pmids: list[str], api_key: str | None = None) -> list[dict]:
    if not pmids:
        return []
    params: dict = {
        "db": "pubmed", "id": ",".join(pmids),
        "retmode": "xml", "rettype": "abstract",
    }
    if api_key:
        params["api_key"] = api_key
    r = requests.get(NCBI_BASE + "efetch.fcgi", params=params, timeout=60)
    r.raise_for_status()
    return _parse_pubmed_xml(r.text)


def _parse_pubmed_xml(xml_text: str) -> list[dict]:
    root = ET.fromstring(xml_text)
    articles = []
    for article in root.findall(".//PubmedArticle"):
        rec: dict = {"Source": "pubmed"}

        elem = article.find(".//PMID")
        pmid = elem.text if elem is not None else ""
        rec["PMID"] = pmid
        rec["UID"]  = pmid

        elem = article.find(".//ArticleTitle")
        rec["Title"] = "".join(elem.itertext()).strip() if elem is not None else ""

        authors = []
        for a in article.findall(".//Author"):
            last = a.findtext("LastName", "")
            fore = a.findtext("ForeName", "")
            if last:
                authors.append(f"{last} {fore}".strip())
        rec["Authors"] = "; ".join(authors)

        elem = article.find(".//Journal/Title")
        rec["Journal"] = elem.text if elem is not None else ""

        elem = article.find(".//PubDate/Year")
        if elem is None:
            elem = article.find(".//PubDate/MedlineDate")
        rec["Year"] = elem.text[:4] if elem is not None else ""

        doi = ""
        for id_elem in article.findall(".//ArticleId"):
            if id_elem.get("IdType") == "doi":
                doi = id_elem.text or ""
                break
        rec["DOI"] = doi

        parts = []
        for txt in article.findall(".//AbstractText"):
            label = txt.get("Label", "")
            text  = "".join(txt.itertext()).strip()
            parts.append(f"{label}: {text}" if label else text)
        rec["Abstract"] = " ".join(parts)

        kws  = ["".join(kw.itertext()) for kw in article.findall(".//Keyword")]
        rec["Keywords"] = "; ".join(kws)

        mesh = [m.text for m in article.findall(".//MeshHeading/DescriptorName") if m.text]
        rec["MeSH_Terms"] = "; ".join(mesh)

        articles.append(rec)
    return articles


# ---------------------------------------------------------------------------
# Europe PMC source
# ---------------------------------------------------------------------------

def search_europepmc(
    query: str, max_results: int, year_from: int | None
) -> list[dict]:
    term = query
    if year_from:
        term = f"({query}) AND (FIRST_PDATE:[{year_from}-01-01 TO 9999-12-31])"
    params = {
        "query": term, "format": "json",
        "pageSize": min(max_results, 1000), "resultType": "core",
    }
    r = requests.get(EPMC_BASE + "search", params=params, timeout=30)
    r.raise_for_status()
    results = r.json().get("resultList", {}).get("result", [])
    return [_parse_epmc_record(rec) for rec in results[:max_results]]


def _parse_epmc_record(r: dict) -> dict:
    pmid = str(r.get("pmid", "")).strip()
    uid  = pmid if pmid else f"EPMC:{r.get('id', '')}"

    kw_obj   = r.get("keywordList") or {}
    kws      = kw_obj.get("keyword", []) if isinstance(kw_obj, dict) else []
    keywords = "; ".join(kws) if isinstance(kws, list) else str(kws)

    mesh_obj = r.get("meshHeadingList") or {}
    mesh_hdg = mesh_obj.get("meshHeading", []) if isinstance(mesh_obj, dict) else []
    mesh = "; ".join(
        m.get("descriptorName", "") for m in mesh_hdg if isinstance(m, dict)
    )

    return {
        "UID": uid, "Source": "europepmc", "PMID": pmid,
        "Title":     r.get("title", "").rstrip("."),
        "Authors":   r.get("authorString", ""),
        "Journal":   r.get("journalTitle", ""),
        "Year":      str(r.get("pubYear", "")),
        "DOI":       r.get("doi", "") or "",
        "Abstract":  r.get("abstractText", "") or "",
        "Keywords":  keywords,
        "MeSH_Terms": mesh,
    }


# ---------------------------------------------------------------------------
# Relevance scoring
# ---------------------------------------------------------------------------

def score_relevance(record: dict, query: str) -> int:
    """0–100 score: title hits (3×), keyword field hits (2×), abstract hits (1×)."""
    terms = [w.lower() for w in query.split() if len(w) > 3]
    if not terms:
        return 0
    title    = record.get("Title", "").lower()
    abstract = record.get("Abstract", "").lower()
    kw_text  = record.get("Keywords", "").lower()
    hits = (
        sum(3 for t in terms if t in title) +
        sum(2 for t in terms if t in kw_text) +
        sum(1 for t in terms if t in abstract)
    )
    return min(100, round(hits / (len(terms) * 6) * 100))


# ---------------------------------------------------------------------------
# AI extraction (optional — requires anthropic + ANTHROPIC_API_KEY)
# ---------------------------------------------------------------------------

_AI_PROMPT = """\
Extract structured information from this biomedical article.
Return ONLY a valid JSON object with these exact keys (use empty string if not found):
- species: semicolon-separated plant or organism scientific names mentioned
- compounds: semicolon-separated chemical compounds or compound classes identified
- activities: semicolon-separated biological activities studied
- key_results: semicolon-separated key quantitative findings (include units, e.g. "IC50=12.3 μg/mL vs DPPH")
- study_type: one of in_vitro | in_vivo | in_silico | clinical | review | mixed

Title: {title}
Abstract: {abstract}

JSON:"""


def ai_extract(record: dict, client, cache: ArticleCache) -> dict:
    uid    = record["UID"]
    cached = cache.get_ai(uid)
    if cached is not None:
        return cached

    abstract = record.get("Abstract", "")
    if not abstract:
        empty = {k: "" for k in ["species", "compounds", "activities", "key_results", "study_type"]}
        cache.put_ai(uid, empty, AI_MODEL)
        return empty

    prompt = _AI_PROMPT.format(
        title=record.get("Title", "")[:300],
        abstract=abstract[:2000],
    )
    try:
        msg = client.messages.create(
            model=AI_MODEL, max_tokens=512,
            messages=[{"role": "user", "content": prompt}],
        )
        text = msg.content[0].text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        result = json.loads(text)
    except Exception:
        result = {k: "" for k in ["species", "compounds", "activities", "key_results", "study_type"]}

    cache.put_ai(uid, result, AI_MODEL)
    return result


def _flatten_ai(extraction: dict) -> dict:
    return {
        "AI_Species":     extraction.get("species", ""),
        "AI_Compounds":   extraction.get("compounds", ""),
        "AI_Activities":  extraction.get("activities", ""),
        "AI_Key_Results": extraction.get("key_results", ""),
        "AI_Study_Type":  extraction.get("study_type", ""),
    }


# ---------------------------------------------------------------------------
# Reference export: BibTeX and RIS
# ---------------------------------------------------------------------------

def _bibtex_authors(authors_str: str) -> str:
    """'Last F; Last2 F2' → 'Last, F and Last2, F2' for BibTeX."""
    parts = [a.strip() for a in authors_str.split(";") if a.strip()]
    result = []
    for part in parts:
        tokens = part.split()
        if len(tokens) >= 2:
            result.append(f"{tokens[0]}, {' '.join(tokens[1:])}")
        else:
            result.append(part)
    return " and ".join(result)


def to_bibtex(df: pd.DataFrame, path: str) -> None:
    lines = []
    for _, rec in df.iterrows():
        pmid = rec.get("PMID") or rec.get("UID", "unknown")
        key  = f"pmid{pmid}"
        lines.append(f"@article{{{key},")
        lines.append(f"  author   = {{{_bibtex_authors(str(rec.get('Authors', '')))}}},")
        lines.append(f"  title    = {{{rec.get('Title', '')}}},")
        lines.append(f"  journal  = {{{rec.get('Journal', '')}}},")
        lines.append(f"  year     = {{{rec.get('Year', '')}}},")
        if rec.get("DOI"):
            lines.append(f"  doi      = {{{rec['DOI']}}},")
        if rec.get("PMID"):
            lines.append(f"  pmid     = {{{rec['PMID']}}},")
        abstract = str(rec.get("Abstract", "")).replace("{", "{{").replace("}", "}}")
        if abstract:
            lines.append(f"  abstract = {{{abstract[:1500]}}},")
        lines.append("}")
        lines.append("")
    Path(path).write_text("\n".join(lines), encoding="utf-8")
    print(f"  BibTeX  → {path} ({len(df)} entries)")


def to_ris(df: pd.DataFrame, path: str) -> None:
    lines = []
    for _, rec in df.iterrows():
        lines.append("TY  - JOUR")
        for author in str(rec.get("Authors", "")).split(";"):
            author = author.strip()
            if not author:
                continue
            tokens = author.split()
            if len(tokens) >= 2:
                author = f"{tokens[0]}, {' '.join(tokens[1:])}"
            lines.append(f"AU  - {author}")
        lines.append(f"TI  - {rec.get('Title', '')}")
        lines.append(f"JO  - {rec.get('Journal', '')}")
        if rec.get("Year"):
            lines.append(f"PY  - {rec['Year']}")
        if rec.get("DOI"):
            lines.append(f"DO  - {rec['DOI']}")
        if rec.get("PMID"):
            lines.append(f"AN  - {rec['PMID']}")
        abstract = str(rec.get("Abstract", ""))
        if abstract:
            lines.append(f"AB  - {abstract[:2000]}")
        for kw in str(rec.get("Keywords", "")).split(";"):
            kw = kw.strip()
            if kw:
                lines.append(f"KW  - {kw}")
        lines.append("ER  - ")
        lines.append("")
    Path(path).write_text("\n".join(lines), encoding="utf-8")
    print(f"  RIS     → {path} ({len(df)} entries)")


# ---------------------------------------------------------------------------
# Excel export with charts
# ---------------------------------------------------------------------------

def _add_charts_sheet(workbook, df: pd.DataFrame) -> None:
    from openpyxl.chart import BarChart, Reference

    ws = workbook.create_sheet("Charts")

    # --- Section 1: Publications per year ---
    year_counts = (
        df["Year"].replace("", pd.NA).dropna()
        .value_counts().sort_index()
    )
    valid_years: dict[int, int] = {}
    for yr, cnt in year_counts.items():
        try:
            valid_years[int(yr)] = int(cnt)
        except (ValueError, TypeError):
            pass
    valid_years = dict(sorted(valid_years.items()))

    ws.cell(1, 1, "Year")
    ws.cell(1, 2, "Records")
    for i, (yr, cnt) in enumerate(valid_years.items(), start=2):
        ws.cell(i, 1, yr)
        ws.cell(i, 2, cnt)
    n_yr = len(valid_years)

    if n_yr >= 2:
        chart1 = BarChart()
        chart1.type   = "col"
        chart1.title  = "Publications per Year"
        chart1.y_axis.title = "Records"
        chart1.x_axis.title = "Year"
        chart1.style  = 10
        chart1.width  = 22
        chart1.height = 12
        chart1.add_data(
            Reference(ws, min_col=2, min_row=1, max_row=1 + n_yr),
            titles_from_data=True,
        )
        chart1.set_categories(
            Reference(ws, min_col=1, min_row=2, max_row=1 + n_yr)
        )
        ws.add_chart(chart1, "D2")

    # --- Section 2: Top 10 journals ---
    j_start = n_yr + 4
    top_journals = (
        df["Journal"].replace("", pd.NA).dropna()
        .value_counts().head(10)
    )
    ws.cell(j_start, 1, "Journal")
    ws.cell(j_start, 2, "Records")
    for i, (journal, cnt) in enumerate(top_journals.items(), start=1):
        ws.cell(j_start + i, 1, journal)
        ws.cell(j_start + i, 2, int(cnt))
    n_j = len(top_journals)

    if n_j >= 2:
        chart2 = BarChart()
        chart2.type   = "bar"
        chart2.title  = "Top 10 Journals"
        chart2.y_axis.title = "Journal"
        chart2.x_axis.title = "Records"
        chart2.style  = 10
        chart2.width  = 22
        chart2.height = 14
        chart2.add_data(
            Reference(ws, min_col=2, min_row=j_start, max_row=j_start + n_j),
            titles_from_data=True,
        )
        chart2.set_categories(
            Reference(ws, min_col=1, min_row=j_start + 1, max_row=j_start + n_j)
        )
        ws.add_chart(chart2, "D22")

    # --- Section 3: Relevance score distribution ---
    s_start = j_start + n_j + 4
    buckets = ["0–19", "20–39", "40–59", "60–79", "80–100"]
    bounds  = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 101)]
    scores  = df["Relevance_Score"].dropna()
    ws.cell(s_start, 1, "Score Range")
    ws.cell(s_start, 2, "Records")
    for i, (label, (lo, hi)) in enumerate(zip(buckets, bounds), start=1):
        cnt = int(((scores >= lo) & (scores < hi)).sum())
        ws.cell(s_start + i, 1, label)
        ws.cell(s_start + i, 2, cnt)

    chart3 = BarChart()
    chart3.type   = "col"
    chart3.title  = "Relevance Score Distribution"
    chart3.y_axis.title = "Records"
    chart3.x_axis.title = "Score Range"
    chart3.style  = 10
    chart3.width  = 16
    chart3.height = 10
    chart3.add_data(
        Reference(ws, min_col=2, min_row=s_start, max_row=s_start + 5),
        titles_from_data=True,
    )
    chart3.set_categories(
        Reference(ws, min_col=1, min_row=s_start + 1, max_row=s_start + 5)
    )
    ws.add_chart(chart3, "D42")

    ws.column_dimensions["A"].width = 45


def export_excel(
    all_df: pd.DataFrame,
    queries: list[str],
    threshold: int,
    output_path: str,
) -> None:
    high_rel_df = all_df[all_df["Relevance_Score"] >= threshold].copy()

    def query_df(q: str) -> pd.DataFrame:
        mask = all_df["Queries"].str.contains(q, regex=False, na=False)
        return all_df[mask].drop(columns=["Queries"]).reset_index(drop=True)

    summary_df = _build_summary_df(all_df, queries)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_df.to_excel(writer,      index=False, sheet_name="All Results")
        high_rel_df.to_excel(writer, index=False, sheet_name="High Relevance")
        query_df(queries[0]).to_excel(writer, index=False, sheet_name=queries[0][:31])
        if len(queries) >= 2:
            query_df(queries[1]).to_excel(writer, index=False, sheet_name=queries[1][:31])
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        wb = writer.book
        _add_charts_sheet(wb, all_df)

        for name in ["All Results", "High Relevance", queries[0][:31]]:
            if name in wb.sheetnames:
                wb[name].freeze_panes = "A2"
        if len(queries) >= 2 and queries[1][:31] in wb.sheetnames:
            wb[queries[1][:31]].freeze_panes = "A2"

    print(
        f"  Excel   → {output_path} "
        f"({len(all_df)} records, {len(high_rel_df)} high-relevance)"
    )


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _build_summary_df(df: pd.DataFrame, queries: list[str]) -> pd.DataFrame:
    rows = []
    for q in queries:
        sub = df[df["Queries"].str.contains(q, regex=False, na=False)]
        rows.append({
            "Query":         q,
            "Records":       len(sub),
            "Sources":       ", ".join(sorted(sub["Source"].unique())) if len(sub) else "",
            "Year_Min":      sub["Year"].replace("", pd.NA).dropna().min() if len(sub) else "",
            "Year_Max":      sub["Year"].replace("", pd.NA).dropna().max() if len(sub) else "",
            "Avg_Relevance": round(sub["Relevance_Score"].mean(), 1) if len(sub) else 0,
        })
    rows.append({
        "Query":         "TOTAL (unique records)",
        "Records":       len(df),
        "Sources":       ", ".join(sorted(df["Source"].unique())) if len(df) else "",
        "Year_Min":      df["Year"].replace("", pd.NA).dropna().min() if len(df) else "",
        "Year_Max":      df["Year"].replace("", pd.NA).dropna().max() if len(df) else "",
        "Avg_Relevance": round(df["Relevance_Score"].mean(), 1) if len(df) else 0,
    })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args    = parse_args()
    queries = args.queries or DEFAULT_QUERIES
    sleep   = 0.12 if args.api_key else 0.4
    cache   = ArticleCache(args.cache_db, bypass=args.no_cache)

    # Validate AI extraction prerequisites
    ai_client = None
    if args.ai_extract:
        if not _ANTHROPIC_OK:
            raise SystemExit("ERROR: --ai-extract requires 'pip install anthropic'")
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            raise SystemExit("ERROR: Set ANTHROPIC_API_KEY env var to use --ai-extract")
        ai_client = _anthropic.Anthropic(api_key=api_key)

    # uid → record dict (deduplicated across sources and queries)
    uid_to_rec:     dict[str, dict]      = {}
    uid_to_queries: dict[str, list[str]] = {}
    doi_to_uid:     dict[str, str]       = {}  # DOI-based cross-source dedup

    def _register(rec: dict, query: str, from_cache: bool = False):
        uid = rec["UID"]
        doi = (rec.get("DOI") or "").lower().strip()

        if doi and doi in doi_to_uid:
            uid = doi_to_uid[doi]
        elif doi:
            doi_to_uid[doi] = uid

        if uid in uid_to_rec:
            uid_to_queries[uid].append(query)
            new_score = score_relevance(uid_to_rec[uid], query)
            if new_score > uid_to_rec[uid]["Relevance_Score"]:
                uid_to_rec[uid]["Relevance_Score"] = new_score
        else:
            rec["Relevance_Score"] = score_relevance(rec, query)
            uid_to_rec[uid]     = rec
            uid_to_queries[uid] = [query]
            if not from_cache:
                cache.put(uid, rec["Source"], rec)

    for query in queries:
        print(f"\n[SEARCH] {query}")

        if args.source in ("pubmed", "both"):
            pmids = search_pubmed(query, args.max_results, args.year_from, args.api_key)
            print(f"  PubMed: {len(pmids)} PMIDs")
            time.sleep(sleep)

            new_pmids: list[str] = []
            for p in pmids:
                cached_rec = cache.get(p)
                if cached_rec:
                    _register(cached_rec, query, from_cache=True)
                else:
                    new_pmids.append(p)

            for i in range(0, len(new_pmids), BATCH_SIZE):
                batch = new_pmids[i : i + BATCH_SIZE]
                print(f"    Fetching {i + 1}–{min(i + BATCH_SIZE, len(new_pmids))} …")
                for rec in fetch_details(batch, args.api_key):
                    _register(rec, query, from_cache=False)
                time.sleep(sleep)

        if args.source in ("europepmc", "both"):
            epmc_records = search_europepmc(query, args.max_results, args.year_from)
            print(f"  EuropePMC: {len(epmc_records)} records")
            for rec in epmc_records:
                cached_rec = cache.get(rec["UID"])
                if cached_rec:
                    _register(cached_rec, query, from_cache=True)
                else:
                    _register(rec, query, from_cache=False)
            time.sleep(sleep)

    if not uid_to_rec:
        print("\nNo records fetched — nothing to export.")
        cache.close()
        return

    # Merge query lists into each record
    for uid, rec in uid_to_rec.items():
        rec["Queries"] = "; ".join(uid_to_queries.get(uid, []))

    # AI extraction
    if ai_client:
        print(f"\n[AI] Extracting from {len(uid_to_rec)} abstracts via {AI_MODEL} …")
        for n, (uid, rec) in enumerate(uid_to_rec.items(), 1):
            extraction = ai_extract(rec, ai_client, cache)
            rec.update(_flatten_ai(extraction))
            if n % 10 == 0:
                print(f"  {n}/{len(uid_to_rec)} processed …")
            time.sleep(0.05)

    cache.close()

    # Build final DataFrame
    columns = COLUMNS_BASE + (COLUMNS_AI if ai_client else [])
    all_df = (
        pd.DataFrame(uid_to_rec.values())
        .reindex(columns=columns, fill_value="")
        .sort_values("Relevance_Score", ascending=False)
        .reset_index(drop=True)
    )

    output_path = args.output or f"pubmed_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_excel(all_df, queries, args.relevance_threshold, output_path)

    if args.export_bibtex:
        to_bibtex(all_df, args.export_bibtex)
    if args.export_ris:
        to_ris(all_df, args.export_ris)

    print(f"\nDone — {len(all_df)} unique records")
    if args.year_from:
        print(f"  Year filter: ≥ {args.year_from}")
    if args.source in ("europepmc", "both"):
        src_counts = all_df["Source"].value_counts()
        for src, cnt in src_counts.items():
            print(f"  {src}: {cnt} records")
    if args.ai_extract:
        print(f"  AI extraction: {AI_MODEL}")


if __name__ == "__main__":
    main()
